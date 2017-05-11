#
# This program uses the examples from the paper "The Intuition 
# Behind Black-Litterman Model  Portfolios", by He and Litterman,
# 1999.  You can find a copy of this  paper at the following url.
#     http://papers.ssrn.com/sol3/papers.cfm?abstract_id=334304
#
#     http://papers.ssrn.com/sol3/papers.cfm?abstract_id=1314585
#
import numpy as np
from scipy import linalg
import urllib2, csv, datetime
import openpyxl#, operator, copy, datetime, pprint
from re import sub
from decimal import Decimal
#from csv import DictWriter
from openpyxl.styles import Font, Alignment, Border, Side
from yahoo_finance import Share
#from datetime import datetime
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
from yahoo_finance import Share
#import datetime as dt
#import pymongo
#from pymongo import MongoClient
from datetime import datetime

# Query index information
print('Downloading index data...')


####################################
###### Begin Index Query ###########
#
##url = "https://www.ishares.com/us/products/239511/ishares-us-healthcare-etf/1449138789749.ajax?fileType=csv&fileName=IYH_holdings&dataType=fund"
#url = "https://www.ishares.com/us/products/239511/ishares-us-healthcare-etf/1467271812596.ajax?fileType=csv&fileName=IYH_holdings&dataType=fund"  # recent change to the download, this updated 8/2016
#
## Seeing about changing index.  SPDR ETF - XLV
#
#response = urllib2.urlopen(url)
#cr = csv.reader(response)
#
#print('                            ...done.')
#
#
#print('Formulating index weights...')
#print('                         ...this may take a minute...')
## Create dictionary with index data
#
##quandl.ApiConfig.api_key = "YOUR_KEY_HERE"    #comeback to this....
#index = 0
#indexCash = 0
#indexCashWeight = 0
#indexData = {}
##healthCareWeight = 0
#totalAssets = 0
#
#for row in cr:
#    if row[0] == 'Fund Holdings as of':
#        reportDate = row[1]
#        endDate = datetime.strptime(reportDate, '%d-%b-%Y')
#        # startDate = endDate - datetime.timedelta(days=7)
#    #elif row[0] == 'Total Net Assets':
#        #totalAssets = Decimal(sub(r'[^\d.]','',row[1]))
#    elif row[0] == 'Ticker':
#        index = 1    
#    elif index == 1 and row[0] != '\xc2\xa0' and str(row[8]).rstrip() == 'Cash and/or Derivatives':
#        indexCash += Decimal(sub(r'[^\d.]','',row[6]))
#    elif index == 1 and row[0] != '\xc2\xa0' and row[8] == 'Health Care':
#        if row[0] != '-':
#            ticker = row[0]
#            #print('Entry {}'.format(ticker)) # for testing
#            shares = float(row[5].replace(',',''))
#            # marketCap = totalAssets = Decimal(sub(r'[^\d.]','',row[6]))
#            size = Decimal(sub(r'[^\d.]','',row[7]))
#            totalAssets += float(size)
#            #print(str(ticker))
#            company = Share(str(ticker))                                                                                               ## re-activate
#            marketCap = company.get_market_cap()
#            if marketCap is not None and marketCap[-1] == 'B':
#                marketCap = float(marketCap[:-1])*1000000
#            elif marketCap is not None and marketCap[-1] == 'M':
#                marketCap = float(marketCap[:-1])*1000
#            else:
#                print('Exception: {} has a non-suffix market cap'.format(ticker))
#            indexData.setdefault(ticker, {'companyName':'', 'sectorWeight':0,
#            'weight':0, 'endPrice':0, 'shares':0, 'industry':'', 'mktCap':0, 'size':0})
#            indexData[ticker]['companyName'] = str(row[1])
#            #indexData[ticker]['weight'] = #float(row[3])
#            indexData[ticker]['endPrice'] = float(row[4])  #endPrice
#            indexData[ticker]['shares'] = shares
#            if marketCap is not None:
#                indexData[ticker]['mktCap'] = float(marketCap) # pull market cap info from Yahoo....
#            else:
#                indexData[ticker]['mktCap'] = 0
#            indexData[ticker]['size'] = float(size)
#            #healthCareWeight += float(row[3])
#            #indexData[ticker]['startPrice'] = float(startPrice)                                                                        ## re-activate
#            #indexData[ticker]['return'] = endPrice/float(startPrice) - 1                                                               ## re-activate
#indexCashWeight = float(indexCash) / totalAssets
#
#for key in indexData:
#    indexData[key]['weight'] = indexData[key]['size']/totalAssets
#
#
###### End Index Query ########
##############################

################################################
##############################
##### Demo Index Query #######

indexData = {}

wb = openpyxl.load_workbook('indexData.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

for row in range(2, sheet.max_row + 1):
    ticker = sheet['A' + str(row)].value
    companyName = sheet['B' + str(row)].value
    endPrice = sheet['C' + str(row)].value
    mktCap = sheet['E' + str(row)].value
    shares = sheet['G' + str(row)].value
    size = sheet['H' + str(row)].value
    weight = sheet['I' + str(row)].value
     
    indexData.setdefault(ticker, {'companyName':'', 'sectorWeight':0,
            'weight':0, 'endPrice':0, 'shares':0, 'industry':'', 'mktCap':0, 'size':0})
    indexData[ticker]['companyName'] = str(companyName)
    indexData[ticker]['endPrice'] = float(endPrice)
    indexData[ticker]['mktCap'] = float(mktCap)
    indexData[ticker]['shares'] = float(shares)
    indexData[ticker]['size'] = float(size)
    indexData[ticker]['weight'] = float(weight)

##### End Demo Index Query####
##############################


# index data does not contain proper industry classification
# look up industry from 'industry library.xlms' saved locally.
industryDict = {}
wb = openpyxl.load_workbook('industry library.xlsx')
sheet = wb.get_sheet_by_name('industries')    
        
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one company.
    ticker = sheet['A' + str(row)].value
    industry = sheet['C' + str(row)].value
        
    industryDict[ticker] = str(industry)

missing = []    
for key in indexData:
    if key in industryDict.keys():
        indexData[key]['industry'] = industryDict[key] 
    else:
        missing.append(key)
if missing: 
    print
    print('Not all companies in the index are classified by industry')
    print('This is probably because a company has been added to the index')
    print('You must enter the file: <industry library.xlsx> and insert the information')
    print('The following companies are missing:')
    print(missing)
    print


print('                            ...done.')

print('Calculating index weights along sector...')

healthCareSvcs = 0
medEquipAndSvcs = 0
pharma = 0
biotech = 0


for key in indexData:
    if (indexData[key]['industry'] == 'Health Care Facilities' or indexData[key]['industry'] == 'Health Care Services' or \
    indexData[key]['industry'] == 'Health Care Management Services' or indexData[key]['industry'] == 'Health Care: Misc'):
        healthCareSvcs += indexData[key]['mktCap']         
    elif (indexData[key]['industry'] == 'Medical & Dental Instruments & Supplies' or indexData[key]['industry'] == 'Medical Equipment' or\
    indexData[key]['industry'] == 'Medical Services'):
        medEquipAndSvcs += indexData[key]['mktCap']
    elif indexData[key]['industry'] == 'Pharmaceuticals':
        pharma += indexData[key]['mktCap']
    elif indexData[key]['industry'] == 'Biotechnology':
        biotech += indexData[key]['mktCap']

for key in indexData:
    if (indexData[key]['industry'] == 'Health Care Facilities' or indexData[key]['industry'] == 'Health Care Services' or \
    indexData[key]['industry'] == 'Health Care Management Services' or indexData[key]['industry'] == 'Health Care: Misc'):
        indexData[key]['sectorWeight'] = indexData[key]['mktCap']/healthCareSvcs         
    elif (indexData[key]['industry'] == 'Medical & Dental Instruments & Supplies' or indexData[key]['industry'] == 'Medical Equipment' or\
    indexData[key]['industry'] == 'Medical Services'):
        indexData[key]['sectorWeight'] = indexData[key]['mktCap'] / medEquipAndSvcs
    elif indexData[key]['industry'] == 'Pharmaceuticals':
        indexData[key]['sectorWeight'] = indexData[key]['mktCap'] / pharma
    elif indexData[key]['industry'] == 'Biotechnology':
        indexData[key]['sectorWeight'] = indexData[key]['mktCap'] / biotech
        
print('                            ...done.')


##############################################
######### Begin Database Query ###############
#
#print('Pulling price data for each stock...')
#weights = []
#stocks = []
#missing = []
#
## Connect to database
#connection = MongoClient('mongodb://localhost:27017/')
#db = connection['health_care_index']    # this establishes connection to DB of said name....
#
## Connect to collection as needed
#data = pd.DataFrame()
#
#for key in indexData:
#    stocks.append(key)
#
#
#for item in stocks:
#    # query database for data....
#    tempData = pd.DataFrame()
#    if item in db.collection_names():        
#        #collection = db[item]
#        tempData = pd.DataFrame(list(db[item].find({"date": {"$gt": datetime(2016, 1, 1,0,0)}})))
#        tempData = tempData.drop(tempData.columns[[0]], axis=1)
#        tempData = tempData.set_index('date')
#        tempData = tempData.sort_index(ascending=True)
#        tempData.columns = [item]
#        #tempData = tempData.fillna(0)
#        data = pd.concat([data, tempData.fillna(0)], axis=1)              
#    else:
#        missing.append(item)
#
#if missing:
#    print('The following are missing: {}'.format(missing))
#
#connection.close()
#print('                            ...done.')
#
#
########## End Database Query ############
##########################################


############################################
######## Begin Price Data Demo Query #######

print('Pulling price data for each stock...')

data = pd.read_excel('data.xlsx')
data = data.set_index(['date'])

print('                            ...done.')

######## End Price Data Demo Query #########
############################################






print('Calculating sector returns...')

sectorDict = {}

rets = np.log(data / data.shift(1))  # dataframe now contains log normal returns
rets = rets.fillna(0)
stocks = rets.columns
for item in stocks:
    rets.loc[:,item] *= indexData[item]['sectorWeight']

sectors = ['Pharmaceuticals', 'Biotechnology', 'Medical Equipment and Services',
    'Health Care Services']
    
df_sectors = pd.DataFrame(0, index = rets.index.values, columns = sectors)
for item in stocks:
    if (indexData[item]['industry'] == 'Health Care Facilities' or indexData[item]['industry'] == 'Health Care Services' or \
    indexData[item]['industry'] == 'Health Care Management Services' or indexData[item]['industry'] == 'Health Care: Misc'):
        df_sectors['Health Care Services'] += rets[item]
    elif (indexData[item]['industry'] == 'Medical & Dental Instruments & Supplies' or indexData[item]['industry'] == 'Medical Equipment' or\
    indexData[item]['industry'] == 'Medical Services'):
        df_sectors['Medical Equipment and Services'] += rets[item]
    elif indexData[item]['industry'] == 'Pharmaceuticals':
        df_sectors['Pharmaceuticals'] += rets[item]
    elif indexData[item]['industry'] == 'Biotechnology':
        df_sectors['Biotechnology'] += rets[item]

df_sectors = df_sectors.replace([np.inf, -np.inf], 0)
mean = df_sectors.mean()*252                  # annualized mean
cov = df_sectors.cov()*252                    # calculates covariance matrix
cor = df_sectors.corr()                       # calculates correlation matrix


#####
#for item in sectors:
#    sectorDict.setdefault(item, {'date':'', 'return':0})
#    for ticker in stocks:
#        if indexData[ticker] == item:
#            


print('                            ...done.')

#
#print('Pulling price data for each stock...')
#output = [mean, cov, cor]
#
#save_xls(output, 'CovOutput.xlsx')
#print('Document saved as <CovOutput.xlsx>')
#print('                            ...done.')


#############################################
#Comment:
#    At this point, mean, and cov have been determined
#    Now look for market capitalization....
#
#############################################

# Read raw market cap numbers from Healthcare Mkt Cap.py
# Calculate totals, determine percentages

weights = []

mktCap = 0

pharmaWt = 0
biotechWt = 0
medEqandSvcsWt = 0
hlthCareSvcsWt = 0

for key in indexData:
    mktCap += indexData[key]['mktCap']

for key in indexData:
    if indexData[key]['industry'] == 'Pharmaceuticals':
        pharmaWt += indexData[key]['mktCap']/mktCap
    elif indexData[key]['industry'] == 'Biotechnology':
        biotechWt += indexData[key]['mktCap']/mktCap
    elif indexData[key]['industry'] == 'Medical & Dental Instruments & Supplies' or indexData[key]['industry'] == 'Medical Equipment' or indexData[key]['industry'] == 'Medical Services':
        medEqandSvcsWt += indexData[key]['mktCap']/mktCap
    elif indexData[key]['industry'] == 'Health Care Facilities' or indexData[key]['industry'] == 'Health Care Services' or indexData[key]['industry'] == 'Health Care Management Services' or indexData[key]['industry'] == 'Health Care: Misc':
        hlthCareSvcsWt += indexData[key]['mktCap']/mktCap

weights.append(pharmaWt)
weights.append(biotechWt)
weights.append(medEqandSvcsWt)
weights.append(hlthCareSvcsWt)



#weights = []
#wb = openpyxl.load_workbook('Health Care Weights.xlsx')
#sheet = wb.get_sheet_by_name('sheet1')
#pharmaCap = float(sheet['B2'].value)
#biotechCap = float(sheet['B3'].value)
#medEquipCap = float(sheet['B5'].value)
#healthCareCap = float(sheet['B4'].value) + float(sheet['B6'].value)
#totalCap = pharmaCap + biotechCap + medEquipCap + healthCareCap


#weights.append(pharmaCap/totalCap)
#weights.append(biotechCap/totalCap)
#weights.append(medEquipCap/totalCap)
#weights.append(healthCareCap/totalCap)



#############################################
##Establish inputs for Black-Litterman ##

weq = np.array(weights)
assets = ['Pharma   ','Biotech  ','Med Eq & Svcs','Hlth Care Svcs']
C = np.array(cor)
Sigma = np.sqrt(np.diag(cov))
refPi = np.array(mean)

##############################################
######## Start of Black-Litterman Area #######
##############################################




# blacklitterman
#   This function performs the Black-Litterman blending of the prior
#   and the views into a new posterior estimate of the returns as
#   described in the paper by He and Litterman.
# Inputs
#   delta  - Risk tolerance from the equilibrium portfolio
#   weq    - Weights of the assets in the equilibrium portfolio
#   sigma  - Prior covariance matrix
#   tau    - Coefficiet of uncertainty in the prior estimate of the mean (pi)
#   P      - Pick matrix for the view(s)
#   Q      - Vector of view returns
#   Omega  - Matrix of variance of the views (diagonal)
# Outputs
#   Er     - Posterior estimate of the mean returns
#   w      - Unconstrained weights computed given the Posterior estimates
#            of the mean and covariance of returns.
#   lambda - A measure of the impact of each view on the posterior estimates.
#
def blacklitterman(delta, weq, sigma, tau, P, Q, Omega):
  # Reverse optimize and back out the equilibrium returns
  # This is formula (12) page 6.
  pi = weq.dot(sigma * delta)
  print(pi)
  # We use tau * sigma many places so just compute it once
  ts = tau * sigma
  # Compute posterior estimate of the mean
  # This is a simplified version of formula (8) on page 4.
  middle = linalg.inv(np.dot(np.dot(P,ts),P.T) + Omega)
  print(middle)
  print(Q-np.expand_dims(np.dot(P,pi.T),axis=1))
  er = np.expand_dims(pi,axis=0).T + np.dot(np.dot(np.dot(ts,P.T),middle),(Q - np.expand_dims(np.dot(P,pi.T),axis=1)))
  # Compute posterior estimate of the uncertainty in the mean
  # This is a simplified and combined version of formulas (9) and (15)
  posteriorSigma = sigma + ts - ts.dot(P.T).dot(middle).dot(P).dot(ts)
  print(posteriorSigma)
  # Compute posterior weights based on uncertainty in mean
  w = er.T.dot(linalg.inv(delta * posteriorSigma)).T
  # Compute lambda value
  # We solve for lambda from formula (17) page 7, rather than formula (18)
  # just because it is less to type, and we've already computed w*.
  lmbda = np.dot(linalg.pinv(P).T,(w.T * (1 + tau) - weq).T)
  return [er, w, lmbda]
# Function to display the results of a black-litterman shrinkage
# Inputs
#   title	- Displayed at top of output
#   assets	- List of assets
#   res		- List of results structures from the bl function
#
def display(title,assets,res):
  # My adjustment, convert set object to a list in order to display properly
  #assets = list(assets)
  #
  er = res[0]
  w = res[1]
  lmbda = res[2]
  print('\n' + title)
  line = 'Sector\t\t'
  for p in range(len(P)):
	line = line + 'P' + str(p) + '\t'
  line = line + 'mu\tw*'
  print(line)

  i = 0;
  for x in assets:
	line = '{0}\t'.format(x)
	for j in range(len(P.T[i])):
		line = line + '{0:.1f}\t'.format(100*P.T[i][j])

	line = line + '{0:.3f}\t{1:.3f}'.format(100*er[i][0],100*w[i][0])
	print(line)
        i = i + 1

  line = 'q\t\t'
  i = 0
  for q in Q:
    line = line + '{0:.2f}\t'.format(100*q[0])
    i = i + 1
  print(line)

  line = 'omega/tau\t'
  i = 0
  for o in Omega:
	line = line + '{0:.5f}\t'.format(o[i]/tau)
	i = i + 1
  print(line)

  line = 'lambda\t\t'
  i = 0
  for l in lmbda:
	line = line + '{0:.5f}\t'.format(l[0])
	i = i + 1
  print(line)




# Take the values from He & Litterman, 1999.
# weq = np.array([0.016,0.022,0.052,0.055,0.116,0.124,0.615])     This is taken care of above, a modification of original
#C = np.array([[ 1.000, 0.488, 0.478, 0.515, 0.439, 0.512, 0.491],
#      [0.488, 1.000, 0.664, 0.655, 0.310, 0.608, 0.779],
#      [0.478, 0.664, 1.000, 0.861, 0.355, 0.783, 0.668],
#      [0.515, 0.655, 0.861, 1.000, 0.354, 0.777, 0.653],
#      [0.439, 0.310, 0.355, 0.354, 1.000, 0.405, 0.306],
#      [0.512, 0.608, 0.783, 0.777, 0.405, 1.000, 0.652],
#      [0.491, 0.779, 0.668, 0.653, 0.306, 0.652, 1.000]])
#Sigma = np.array([0.160, 0.203, 0.248, 0.271, 0.210, 0.200, 0.187])
#refPi = np.array([0.039, 0.069, 0.084, 0.090, 0.043, 0.068, 0.076])
#assets= ['Australia','Canada   ','France   ','Germany  ','Japan    ','UK       ','USA      ']  # changed this to a list ([]) from a set ({})



# Equilibrium covariance matrix
V = np.multiply(np.outer(Sigma,Sigma), C)
#print(V)

# Risk aversion of the market 
delta = 2.5

# Coefficient of uncertainty in the prior estimate of the mean
# from footnote (8) on page 11
tau = 0.05
tauV = tau * V

# Define view 1
# Germany will outperform the other European markets by 5%
# Market cap weight the P matrix
# Results should match Table 4, Page 21
P1 = np.array([-.5, .5, .5, -.5])
Q1 = np.array([0.0405])
P=np.array([P1])
Q=np.array([Q1]);
Omega = np.dot(np.dot(P,tauV),P.T) * np.eye(Q.shape[0])
res = blacklitterman(delta, weq, V, tau, P, Q, Omega)
display('View 1',assets,res)

# Define view 2
# Canadian Equities will outperform US equities by 3%
# Market cap weight the P matrix
# Results should match Table 5, Page 22
P2 = np.array([0, 0, -1, 1])
Q2 = np.array([0.03])
P=np.array([P1,P2])
Q=np.array([Q1,Q2]);
Omega = np.dot(np.dot(P,tauV),P.T) * np.eye(Q.shape[0])
res = blacklitterman(delta, weq, V, tau, P, Q, Omega)
display('View 1 + 2', assets, res)


###################################
# Calculating Equilibrium weights #

    
      